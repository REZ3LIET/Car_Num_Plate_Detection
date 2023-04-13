import pytesseract as pt
import cv2
import time
import keyboard
from openpyxl import*

pt.pytesseract.tesseract_cmd = r'loctation\Tesseract\Tesseract-OCR\tesseract.exe'

wb = load_workbook("Register_1.xlsx")
ws = wb['Sheet1']

m1 = 0
m2 = 0
text=''
l_numPlate = []
l_inTime = []
l_outTime = []
def numPlate():
    n = []

    # To load the image to code from its directory
    img = cv2.imread('Pics\\002.jpg')
    cv2.imshow("Original", img)

    # Converting color format to grayscale to suit Tesseract
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    gray = cv2.bilateralFilter(gray, 11, 17, 17)
    # cv2.imshow("gray", gray)

    edged = cv2.Canny(gray, 170, 200)
    # cv2.imshow("Canny", edged)

    cnt, new = cv2.findContours(edged.copy(), cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

    img1 = img.copy()
    cv2.drawContours(img1, cnt, -1, (0, 0, 255), 3)
    # cv2.imshow("Canny after contours", img1)

    cnt = sorted(cnt, key=cv2.contourArea, reverse=True)[:5]
    numPlate = None

    img2 = img.copy()
    cv2.drawContours(img2, cnt, -1, (255, 0, 0), 3)
    # cv2.imshow("Top 5", img2)

    count = 0
    name = 1  # Cropped image name

    for i in cnt:
        peri = cv2.arcLength(i, True)
        approx = cv2.approxPolyDP(i, 0.18 * peri, True)
        if (len(approx) == 4):
            numPlate = approx
            # cropping rectangle
            x, y, w, h = cv2.boundingRect(i)
            crop = img[y:y + h, x:x + w]
            # print(4)
            cv2.imwrite(str(name) + '.png', crop)
            name += 1
            break
        elif (len(approx) == 2):
            numPlate = approx
            # cropping rectangle
            x, y, w, h = cv2.boundingRect(i)
            crop = img[y:y + h, x:x + w]
            # print(2)
            cv2.imwrite(str(name) + '.png', crop)
            name += 1
            break

    cropImg = '1.png'
    textImg = cv2.imread(cropImg)

    cropGray = cv2.cvtColor(textImg, cv2.COLOR_BGR2GRAY)
    cropGray = cv2.bilateralFilter(cropGray, 11, 17, 17)
    cv2.imshow("Cropped", cropGray)

    # Converting Image to Text using Pytesseract
    text = pt.image_to_string(cropGray)

    # Filtering blank inputs and adding to list N
    if text != '':
        n.append(text)

    # Filtering elements of the list
    for i in n:
        if len(i) == 10:
            text = i

    text = pt.image_to_string(cropImg)

    text = text[0:10]

    print("Number on Plate is: " + text)
    l_numPlate.append(str(text))

    cv2.waitKey(0)

def dateTime():
    t = time.localtime()
    curTime = time.strftime("%H:%M:%S", t)
    print("IN TIME: ", curTime)
    l_inTime.append(curTime)

def dateTimeOut():
    t = time.localtime()
    curTime = time.strftime("%H:%M:%S", t)
    print("OUT TIME: ", curTime)
    l_outTime.append(curTime)

def excel():


    l = [l_numPlate, l_inTime]

    rowNum = ws.max_row
    ws.cell(row=rowNum + 1, column=1, value=rowNum)
    ws.cell(row=rowNum + 1, column=2, value=l_numPlate[0])
    ws.cell(row=rowNum + 1, column=3, value=l_inTime[0])
    wb.save("Register_1.xlsx")

def run():
    numPlate()
    while True:

        print("Is displayed value of Number Plate correct?\n\nPress Y for YES else N for NO or E to EXIT.")
        if keyboard.read_key() == 'y':
            scan()
            break

        if keyboard.read_key() == 'e':
            print("You've Exited!!")
            break

        elif keyboard.read_key() == 'n':
            run()

        else:
            pass

def scan():
    i = 1
    rowNum = ws.max_row
    while i < rowNum + 1:
        if(l_numPlate[0] != ws.cell(column=2, row=i).value):
            if(i == rowNum):
                dateTime()
                excel()
                break
            else:
                i+=1

        else:
            dateTimeOut()
            ws.cell(row= i, column= 4, value= l_outTime[0])
            ws.cell(row= i, column= 2, value= l_numPlate[0]+" [OUT]")
            amount(i)
            wb.save("Register_1.xlsx")
            return i
            break

def amount(x):

    inT = ws.cell(column=3, row=x).value
    outT = ws.cell(column=4, row=x).value

    inT_n = inT.split(":")
    outT_n = outT.split(":")

    tIn = int(inT_n[0]) * 3600 + int(inT_n[1]) * 60 + int(inT_n[2])
    tOut = int(outT_n[0]) * 3600 + int(outT_n[1]) * 60 + int(outT_n[2])

    amt_h = ((tOut - tIn)/3600)
    amt =  (amt_h*100)//1
    ws.cell(row=x, column=5, value=amt)
    wb.save("Register_1.xlsx")

run()